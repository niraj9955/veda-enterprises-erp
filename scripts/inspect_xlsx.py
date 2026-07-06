import openpyxl
import json

wb = openpyxl.load_workbook('/home/z/my-project/upload/Production.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Print first 15 rows
    print("\n--- First 15 rows ---")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            break
        print(f"Row {i}: {row}")
    
    # Print header row detailed
    print("\n--- Header row cells ---")
    for cell in ws[1]:
        if cell.value is not None:
            print(f"  Col {cell.column_letter}: '{cell.value}' (type: {type(cell.value).__name__})")
